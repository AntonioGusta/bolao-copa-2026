import os
import io
import datetime as dt
import streamlit as st
import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ==============================================================================
# CONFIGURAÇÃO DE ACESSO
# ==============================================================================
ID_PASTA_DRIVE = "1YXz96rcj7IgthzYHK6uyei_KrC5q9cqb"

def obter_serviço_drive():
    creds = Credentials(
        token=None,
        refresh_token=st.secrets["gcp_oauth"]["refresh_token"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=st.secrets["gcp_oauth"]["client_id"],
        client_secret=st.secrets["gcp_oauth"]["client_secret"]
    )
    if not creds.valid:
        creds.refresh(Request())
    return build('drive', 'v3', credentials=creds)

@st.cache_data(ttl=dt.timedelta(hours=24), show_spinner=False)
def carregar_palpites_em_cache(file_id):
    service = obter_serviço_drive()
    req = service.files().get_media(fileId=file_id)
    bytes_io = io.BytesIO()
    baixador = MediaIoBaseDownload(bytes_io, req)
    while not baixador.next_chunk()[1]: pass
    bytes_io.seek(0)
    
    wb = openpyxl.load_workbook(bytes_io, data_only=True)
    palpites = {'fase1': {}, 'fase2': {}}
    
    # Carrega a Fase de Grupos
    if 'FASE 1' in wb.sheetnames:
        ws1 = wb['FASE 1']
        for r in range(3, 75):
            palpites['fase1'][r] = (ws1[f'G{r}'].value, ws1[f'I{r}'].value)
            
    # Carrega o Mata-Mata
    nome_aba_fase2 = 'FASE 2, 3, 4, SEMI & FINAL'
    if nome_aba_fase2 in wb.sheetnames:
        ws2 = wb[nome_aba_fase2]
        for r in range(3, 100):
            palpites['fase2'][r] = (ws2[f'E{r}'].value, ws2[f'G{r}'].value)
            
    wb.close()
    return palpites

def calcular_pontos(g_m_real, g_v_real, g_m_palpite, g_v_palpite):
    if g_m_real is None or g_v_real is None or g_m_palpite is None or g_v_palpite is None: return 0
    try:
        g_m_real, g_v_real = int(g_m_real), int(g_v_real)
        g_m_palpite, g_v_palpite = int(g_m_palpite), int(g_v_palpite)
    except (ValueError, TypeError): return 0
    
    res_real = 1 if g_m_real > g_v_real else (-1 if g_m_real < g_v_real else 0)
    res_palpite = 1 if g_m_palpite > g_v_palpite else (-1 if g_m_palpite < g_v_palpite else 0)
    if res_real != res_palpite: return 0
    if g_m_real == g_m_palpite and g_v_real == g_v_palpite: return 7
    if res_real == 0: return 3
    if g_m_real == g_m_palpite or g_v_real == g_v_palpite: return 4
    return 2

def gerar_tabela_html(participantes, titulo, subtitulo, cor_destaque=None, posicoes_destaque=None):
    if posicoes_destaque is None: posicoes_destaque = []
    
    html = f"<div style='background: #0F172A; padding: 15px; border-radius: 8px; box-shadow: 0 8px 24px rgba(0,0,0,.35); text-align: center; margin-bottom: 20px; font-family: sans-serif;'>"
    html += f"<h4 style='margin-bottom: 0px; color: #F8FAFC;'>{titulo}</h4>"
    html += f"<p style='margin-top: 0px; font-size: 14px; font-style: italic; color: #CBD5E1;'>{subtitulo}</p>"
    
    html += "<table style='width: 100%; border-collapse: collapse; text-align: center; font-size: 13px; color: #F8FAFC; margin-top: 15px;'>"
    html += "<thead style='background-color: #111827; border-bottom: 2px solid #334155;'>"
    # --- CABEÇALHO ATUALIZADO COM QUARTAS ---
    html += "<tr><th style='padding: 8px;'>Pos</th><th style='padding: 8px;'>Var</th><th style='text-align: left; padding: 8px;'>Nome</th><th style='padding: 8px; color:#94A3B8;'>Grupos</th><th style='padding: 8px; color:#94A3B8;'>Pré-Oitavas</th><th style='padding: 8px; color:#94A3B8;'>Oitavas</th><th style='padding: 8px; color:#94A3B8;'>Quartas</th><th style='padding: 8px; font-size: 15px;'>Total</th><th style='padding: 8px; color:#64748B;'>Dif</th></tr>"
    html += "</thead><tbody>"

    for p in participantes:
        pos_str = p['posicao'].replace("º Lugar", "").strip()
        is_highlight = (pos_str.isdigit() and int(pos_str) in posicoes_destaque)
        bg_color = cor_destaque if is_highlight else "#1E293B"
        
        html += f"<tr style='background-color: {bg_color}; border-bottom: 1px solid #334155;'>"
        html += f"<td style='padding: 6px; font-weight: bold;'>{p['posicao'].replace(' Lugar', '')}</td>"
        html += f"<td style='padding: 6px; font-size: 13px;'>{p.get('var_html', '➖')}</td>"
        html += f"<td style='text-align: left; padding: 6px;'>{p.get('nome_exibicao', p['nome'])}</td>"
        
        # --- COLUNAS DE PONTOS ---
        html += f"<td style='padding: 6px; color:#CBD5E1;'>{p.get('pts_grupos', 0)}</td>"
        html += f"<td style='padding: 6px; color:#CBD5E1;'>{p.get('pts_pre_oitavas', 0)}</td>"
        html += f"<td style='padding: 6px; color:#CBD5E1;'>{p.get('pts_oitavas', 0)}</td>"
        html += f"<td style='padding: 6px; color:#CBD5E1;'>{p.get('pts_quartas', 0)}</td>"
        html += f"<td style='padding: 6px; font-weight: bold; font-size: 15px; color:#F59E0B;'>{p['pontos']}</td>"
        
        html += f"<td style='padding: 6px; font-size: 12px; color:#64748B;'>{p.get('dif', '-')}</td>"
        html += "</tr>"

    html += "</tbody></table></div>"
    return html

# ==============================================================================
# CONFIGURAÇÃO VISUAL E GESTÃO DE ESTADO
# ==============================================================================
st.set_page_config(page_title="Bolão Copa 2026 (THE)", page_icon="⚽", layout="wide")

st.markdown("""
<style>
button[kind="primary"] {
    background-color: #38BDF8 !important; 
    color: #0F172A !important;
    border-color: #38BDF8 !important;
    font-weight: bold;
}
button[kind="primary"]:hover {
    background-color: #0EA5E9 !important; 
    color: #ffffff !important;
}
</style>
""", unsafe_allow_html=True)

if 'ranking_processado' not in st.session_state: st.session_state.ranking_processado = []
if 'resumo_ontem' not in st.session_state: st.session_state.resumo_ontem = {}

st.title("🏆 Apuração do Bolão da Copa 2026 (THE) - V2")
st.write("Clique no botão abaixo para processar os palpites e atualizar o ranking em tempo real.")

if st.button("🚀 Atualizar Classificação", type="primary"):
    with st.spinner("Sincronizando com o Drive..."):
        try:
            service = obter_serviço_drive()
            query = f"'{ID_PASTA_DRIVE}' in parents and trashed = false"
            resultados = service.files().list(q=query, fields="files(id, name)").execute()
            mapa_arquivos = {arq['name']: arq['id'] for arq in resultados.get('files', [])}
            
            if 'Arquivo_de_controle.xlsx' not in mapa_arquivos:
                st.error("Erro: 'Arquivo_de_controle.xlsx' não encontrado na pasta do Drive.")
                st.stop()

            id_controle = mapa_arquivos['Arquivo_de_controle.xlsx']
            requisicao = service.files().get_media(fileId=id_controle)
            bytes_controle = io.BytesIO()
            baixador = MediaIoBaseDownload(bytes_controle, requisicao)
            while not baixador.next_chunk()[1]: pass
                
            bytes_controle.seek(0)
            wb_leitura = openpyxl.load_workbook(bytes_controle, data_only=True)
            
            fuso_br = dt.timezone(dt.timedelta(hours=-3))
            hoje_date = dt.datetime.now(fuso_br).date()
            ontem_date = hoje_date - dt.timedelta(days=1)
            anteontem_date = hoje_date - dt.timedelta(days=2)
            
            # --- COLETAR GABARITO FASE 1 ---
            ws_fase1_leitura = wb_leitura['FASE 1']
            resultados_reais_f1 = {}
            datas_jogos_f1 = {}
            for r in range(3, 75):
                resultados_reais_f1[r] = (ws_fase1_leitura[f'G{r}'].value, ws_fase1_leitura[f'I{r}'].value)
                
                val_data = ws_fase1_leitura[f'C{r}'].value
                d_jogo = None
                if isinstance(val_data, dt.datetime): d_jogo = val_data.date()
                elif val_data:
                    txt = str(val_data).strip()
                    if "/" in txt:
                        try: d_jogo = dt.date(2026, int(txt.split('/')[1]), int(txt.split('/')[0]))
                        except: pass
                datas_jogos_f1[r] = d_jogo

            # --- COLETAR GABARITO MATA-MATA (FASE 2) ---
            nome_aba_fase2 = 'FASE 2, 3, 4, SEMI & FINAL'
            resultados_reais_f2 = {}
            datas_jogos_f2 = {}
            if nome_aba_fase2 in wb_leitura.sheetnames:
                ws_fase2_leitura = wb_leitura[nome_aba_fase2]
                for r in range(3, 100):
                    nome_casa = ws_fase2_leitura[f'D{r}'].value
                    if nome_casa is not None and str(nome_casa).strip() != "":
                        resultados_reais_f2[r] = (ws_fase2_leitura[f'E{r}'].value, ws_fase2_leitura[f'G{r}'].value)
                        
                        val_data = ws_fase2_leitura[f'A{r}'].value
                        d_jogo = None
                        if isinstance(val_data, dt.datetime): d_jogo = val_data.date()
                        elif val_data:
                            txt = str(val_data).strip()
                            if "/" in txt:
                                try: d_jogo = dt.date(2026, int(txt.split('/')[1]), int(txt.split('/')[0]))
                                except: pass
                        datas_jogos_f2[r] = d_jogo

            ws_tabela_leitura = wb_leitura['TABELA']
            lista_ranking = []
            lista_ontem = []
            lista_anteontem = []
            
            # PROCESSAR PLANILHAS
            for row_tabela in range(2, 101):
                celula_nome = ws_tabela_leitura[f'A{row_tabela}'].value
                if celula_nome is None or str(celula_nome).strip() == "" or str(celula_nome).strip().lower() == "participante": continue
                    
                nome = str(celula_nome).strip()
                arquivo_palpite = f"{nome}.xlsx"
                
                if arquivo_palpite not in mapa_arquivos:
                    lista_ranking.append({'nome': nome, 'pontos': 0, 'pts_grupos': 0, 'pts_pre_oitavas': 0, 'pts_oitavas': 0, 'pts_quartas': 0})
                    continue
                    
                id_part = mapa_arquivos[arquivo_palpite]
                palpites = carregar_palpites_em_cache(id_part)
                
                pts_grupos = 0
                pts_pre_oitavas = 0
                pts_oitavas = 0
                pts_quartas = 0
                pts_ont = 0
                pts_ant = 0
                
                # Pontuação Fase 1
                for r, (g_m_real, g_v_real) in resultados_reais_f1.items():
                    if g_m_real is None or g_v_real is None: continue
                    
                    p_casa, p_fora = palpites['fase1'].get(r, (None, None))
                    pts = calcular_pontos(g_m_real, g_v_real, p_casa, p_fora)
                    pts_grupos += pts
                    
                    d_jogo = datas_jogos_f1.get(r)
                    if d_jogo:
                        if d_jogo <= ontem_date: pts_ont += pts
                        if d_jogo <= anteontem_date: pts_ant += pts
                    else:
                        pts_ont += pts
                        pts_ant += pts

                # Pontuação Mata-Mata
                for r, (g_m_real, g_v_real) in resultados_reais_f2.items():
                    if g_m_real is None or g_v_real is None: continue
                    
                    p_casa, p_fora = palpites['fase2'].get(r, (None, None))
                    pts = calcular_pontos(g_m_real, g_v_real, p_casa, p_fora)
                    
                    # FISCAL DE TRÂNSITO: Separando as pontuações pelas linhas
                    if 3 <= r <= 18:
                        pts_pre_oitavas += pts
                    elif 22 <= r <= 29:
                        pts_oitavas += pts
                    elif 33 <= r <= 36:
                        pts_quartas += pts
                    
                    d_jogo = datas_jogos_f2.get(r)
                    if d_jogo:
                        if d_jogo <= ontem_date: pts_ont += pts
                        if d_jogo <= anteontem_date: pts_ant += pts
                    else:
                        pts_ont += pts
                        pts_ant += pts
                        
                pts_total_live = pts_grupos + pts_pre_oitavas + pts_oitavas + pts_quartas
                
                lista_ranking.append({
                    'nome': nome, 
                    'pontos': pts_total_live, 
                    'pts_grupos': pts_grupos,
                    'pts_pre_oitavas': pts_pre_oitavas,
                    'pts_oitavas': pts_oitavas,
                    'pts_quartas': pts_quartas
                })
                
                lista_ontem.append({'nome': nome, 'pontos': pts_ont})
                lista_anteontem.append({'nome': nome, 'pontos': pts_ant})
                
            wb_leitura.close()
            
            # ==================================================================
            # CALCULAR RANKS PASSADOS ON THE FLY
            # ==================================================================
            def ranquear_passado(lista):
                lista.sort(key=lambda x: (-x['pontos'], x['nome'].lower()))
                pos_dict = {}
                posicao = 1
                ultimos_pts = None
                for idx, p in enumerate(lista):
                    if p['pontos'] != ultimos_pts: posicao = idx + 1
                    ultimos_pts = p['pontos']
                    pos_dict[p['nome']] = posicao
                return pos_dict

            ranks_ontem = ranquear_passado(lista_ontem)
            ranks_anteontem = ranquear_passado(lista_anteontem)

            var_ontem_cards = {}
            for n, pos_ont in ranks_ontem.items():
                pos_ant = ranks_anteontem.get(n, pos_ont)
                var_ontem_cards[n] = pos_ant - pos_ont
                
            m_subida = max(var_ontem_cards.values()) if var_ontem_cards else 0
            m_queda = min(var_ontem_cards.values()) if var_ontem_cards else 0
            
            h_sub = [n for n, v in var_ontem_cards.items() if v == m_subida] if m_subida > 0 else []
            v_que = [n for n, v in var_ontem_cards.items() if v == m_queda] if m_queda < 0 else []
            
            st.session_state.resumo_ontem = {
                "data_str": f"{ontem_date.day:02d}/{ontem_date.month:02d}",
                "maior_subida": {"nomes": h_sub, "valor": m_subida},
                "maior_queda": {"nomes": v_que, "valor": abs(m_queda)}
            }

            # RANQUEAR E FORMATAR LISTA LIVE
            lista_ranking.sort(key=lambda x: (-x['pontos'], x['nome'].lower()))
            
            posicao_atual = 1
            pontos_anteriores = None
            
            for idx, participante in enumerate(lista_ranking):
                if participante['pontos'] != pontos_anteriores: posicao_atual = idx + 1
                pontos_anteriores = participante['pontos']
                participante['posicao'] = f"{posicao_atual}º Lugar"
                
                emoji = ""
                if posicao_atual == 1: emoji = "🥇 "
                elif posicao_atual == 2: emoji = "🥈 "
                elif posicao_atual == 3: emoji = "🥉 "
                participante['nome_exibicao'] = f"{emoji}{participante['nome']}"
                
                if idx < len(lista_ranking) - 1:
                    dif = participante['pontos'] - lista_ranking[idx+1]['pontos']
                    participante['dif'] = f"+{dif}" if dif > 0 else "="
                else: participante['dif'] = "-"

                pos_antiga = ranks_ontem.get(participante['nome'], posicao_atual)
                variacao = pos_antiga - posicao_atual
                
                if variacao > 0: participante['var_html'] = f"<span style='color: #22C55E;'>▲ {variacao}</span>"
                elif variacao < 0: participante['var_html'] = f"<span style='color: #EF4444;'>▼ {abs(variacao)}</span>"
                else: participante['var_html'] = "<span style='color: #94A3B8;'>➖</span>"

            # GRAVAR EXCEL (PROTEGIDO COMO AMBIENTE DE TESTES)
            bytes_controle.seek(0)
            wb_gravar = openpyxl.load_workbook(bytes_controle, data_only=False)
            if 'Dados Pessoais' in wb_gravar.sheetnames: wb_gravar.remove(wb_gravar['Dados Pessoais'])
            ws_tabela_gravar = wb_gravar['TABELA']
            for r in range(1, 101): ws_tabela_gravar[f'A{r}'].value, ws_tabela_gravar[f'B{r}'].value, ws_tabela_gravar[f'C{r}'].value = None, None, None
            ws_tabela_gravar['A1'].value, ws_tabela_gravar['B1'].value, ws_tabela_gravar['C1'].value = "Participante", "Pontos", "Posição"
            
            for idx, p in enumerate(lista_ranking):
                linha_atual = 2 + idx
                ws_tabela_gravar[f'A{linha_atual}'].value = p['nome']
                ws_tabela_gravar[f'B{linha_atual}'].value = p['pontos']
                ws_tabela_gravar[f'C{linha_atual}'].value = p['posicao']
                
            st.session_state.ranking_processado = lista_ranking.copy()
            
            arquivo_saida_bytes = io.BytesIO()
            wb_gravar.save(arquivo_saida_bytes)
            arquivo_saida_bytes.seek(0)
            wb_gravar.close()
            
            nome_saida = 'BOLÃO DA COPA DO MUNDO 2026-TESTE.xlsx'
            media_upload = MediaIoBaseUpload(arquivo_saida_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
            if nome_saida in mapa_arquivos: service.files().update(fileId=mapa_arquivos[nome_saida], media_body=media_upload).execute()
            else: service.files().create(body={'name': nome_saida, 'parents': [ID_PASTA_DRIVE]}, media_body=media_upload).execute()
            
            st.markdown("<div style='background: #064E3B; color: #D1FAE5; border-left: 4px solid #10B981; padding: 12px; border-radius: 4px; margin-bottom: 20px;'><strong>Tabela atualizada com sucesso no ambiente de testes!</strong></div>", unsafe_allow_html=True)
            
        except Exception as e: st.error(f"Ocorreu um erro no processamento: {e}")

st.divider()

# ==============================================================================
# NAVEGAÇÃO SEGURA (CONTORNO DO BUG DE ABAS DO STREAMLIT)
# ==============================================================================
aba_selecionada = st.radio(
    "Navegação:",
    ["🏆 Classificação e Resenha", "👀 Espiar Palpites da Rodada"],
    horizontal=True,
    label_visibility="collapsed"
)

# --- TELA 1: RANKING ESTILIZADO ---
if aba_selecionada == "🏆 Classificação e Resenha":
    if not st.session_state.ranking_processado: 
        st.info("👆 Clique no botão azul lá em cima para buscar os dados.")
    else:
        dados = st.session_state.ranking_processado
        resumo = st.session_state.get('resumo_ontem', {})
        N = len(dados)
        
        # --- CARDS DE RESUMO (LÍDER, SUBIDA, QUEDA) ---
        if dados:
            lider = dados[0]
            col_c1, col_c2, col_c3 = st.columns(3)
            
            with col_c1:
                st.markdown(f"""
                <div style="background: linear-gradient(90deg, #1E293B 0%, #0F172A 100%); border-left: 4px solid #F59E0B; padding: 15px; border-radius: 6px; margin-bottom: 25px; min-height: 105px;">
                    <h5 style="color: #FBBF24; margin: 0 0 5px 0;">🏆 Líder Geral</h5>
                    <p style="color: #F8FAFC; margin: 0; font-size: 14px;"><strong>{lider['nome']}</strong> segue no topo com <strong>{lider['pontos']} pts</strong>!</p>
                </div>
                """, unsafe_allow_html=True)
                
            with col_c2:
                val_sub = resumo.get('maior_subida', {}).get('valor', 0)
                data_str = resumo.get('data_str', 'Ontem')
                nomes_sub = resumo.get('maior_subida', {}).get('nomes', [])
                
                if val_sub > 0 and nomes_sub:
                    n_str = ", ".join(nomes_sub)
                    st.markdown(f"""
                    <div style="background: linear-gradient(90deg, #1E293B 0%, #0F172A 100%); border-left: 4px solid #22C55E; padding: 15px; border-radius: 6px; margin-bottom: 25px; min-height: 105px;">
                        <h5 style="color: #4ADE80; margin: 0 0 5px 0;">🚀 Maior Subida ({data_str})</h5>
                        <p style="color: #F8FAFC; margin: 0; font-size: 14px;"><strong>{n_str}</strong> saltou <strong>+{val_sub} posições</strong></p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background: #0F172A; border-left: 4px solid #334155; padding: 15px; border-radius: 6px; margin-bottom: 25px; min-height: 105px;">
                        <h5 style="color: #94A3B8; margin: 0 0 5px 0;">🚀 Maior Subida ({data_str})</h5>
                        <p style="color: #64748B; margin: 0; font-size: 14px;">Nenhuma alteração na tabela.</p>
                    </div>
                    """, unsafe_allow_html=True)

            with col_c3:
                val_que = resumo.get('maior_queda', {}).get('valor', 0)
                nomes_que = resumo.get('maior_queda', {}).get('nomes', [])
                
                if val_que > 0 and nomes_que:
                    n_str = ", ".join(nomes_que)
                    st.markdown(f"""
                    <div style="background: linear-gradient(90deg, #1E293B 0%, #0F172A 100%); border-left: 4px solid #EF4444; padding: 15px; border-radius: 6px; margin-bottom: 25px; min-height: 105px;">
                        <h5 style="color: #F87171; margin: 0 0 5px 0;">📉 Maior Queda ({data_str})</h5>
                        <p style="color: #F8FAFC; margin: 0; font-size: 14px;"><strong>{n_str}</strong> caiu <strong>-{val_que} posições</strong></p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background: #0F172A; border-left: 4px solid #334155; padding: 15px; border-radius: 6px; margin-bottom: 25px; min-height: 105px;">
                        <h5 style="color: #94A3B8; margin: 0 0 5px 0;">📉 Maior Queda ({data_str})</h5>
                        <p style="color: #64748B; margin: 0; font-size: 14px;">Nenhuma alteração na tabela.</p>
                    </div>
                    """, unsafe_allow_html=True)

        idx_prof = 5 if N >= 5 else N
        while idx_prof < N and dados[idx_prof]['pontos'] == dados[idx_prof - 1]['pontos']: idx_prof += 1
            
        idx_lant = max(N - 3, idx_prof) 
        while idx_lant > idx_prof and dados[idx_lant - 1]['pontos'] == dados[idx_lant]['pontos']: idx_lant -= 1
            
        alvo_amad = idx_prof + 10
        idx_amad = min(alvo_amad, idx_lant) 
        while idx_amad < idx_lant and dados[idx_amad]['pontos'] == dados[idx_amad - 1]['pontos']: idx_amad += 1
            
        profissionais, amadores = dados[:idx_prof], dados[idx_prof:idx_amad]
        peladeiros, lanterna = dados[idx_amad:idx_lant], dados[idx_lant:]
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if profissionais: st.markdown(gerar_tabela_html(profissionais, "Profissionais", "- Elite do Pitaco -", "#166534", [1, 2, 3, 4, 5]), unsafe_allow_html=True)
        with col2:
            if amadores: st.markdown(gerar_tabela_html(amadores, "Amadores", "- Os que ainda sonham -"), unsafe_allow_html=True)
        with col3:
            if peladeiros: st.markdown(gerar_tabela_html(peladeiros, "Peladeiros", "- Especialistas em Errar -"), unsafe_allow_html=True)
            
        st.write("---")
        col_vazia1, col_lanterna, col_vazia2 = st.columns([1, 2, 1])
        with col_lanterna:
            if lanterna:
                posicoes_lanterna = [int(p['posicao'].replace("º Lugar", "").strip()) for p in lanterna]
                st.markdown(gerar_tabela_html(lanterna, "Prêmio Espírito Coletivo", "- Bastava Apostar ao Contrário -", "#991B1B", posicoes_lanterna), unsafe_allow_html=True)

# --- TELA 2: PALPITES DO DIA ---
elif aba_selecionada == "👀 Espiar Palpites da Rodada":
    st.subheader("👀 Espiar Palpites da Rodada")
    fuso_br = dt.timezone(dt.timedelta(hours=-3))
    hoje = dt.datetime.now(fuso_br)
    data_padrao = f"{hoje.day}/{hoje.month}"
    
    if 'busca_ativa' not in st.session_state:
        st.session_state.busca_ativa = False
    if 'data_pesquisada' not in st.session_state:
        st.session_state.data_pesquisada = ""

    data_pesquisa = st.text_input("📅 Digite a data dos jogos que deseja ver (Ex: 12/6):", value=data_padrao)
    
    if st.button("🔍 Buscar Rodada", type="secondary"):
        st.session_state.busca_ativa = True
        st.session_state.data_pesquisada = data_pesquisa

    if st.session_state.busca_ativa:
        with st.spinner("Analisando palpites..."):
            try:
                from collections import Counter
                service = obter_serviço_drive()
                
                pesq_bruta = st.session_state.data_pesquisada.strip().lower()
                modo_grafico = False
                
                if "-graficos" in pesq_bruta:
                    modo_grafico = True
                    pesq_bruta = pesq_bruta.replace("-graficos", "").strip()
                
                dia_pesq, mes_pesq = None, None
                if "/" in pesq_bruta:
                    try: 
                        dia_pesq = int(pesq_bruta.split('/')[0])
                        mes_pesq = int(pesq_bruta.split('/')[1])
                        pesq_limpa = f"{dia_pesq}/{mes_pesq}"
                    except: pesq_limpa = pesq_bruta
                else: pesq_limpa = pesq_bruta
                
                # --- Lógica da Data: Fase 1 (<= 27/6) vs Mata-Mata (>= 28/6) ---
                is_mata_mata = False
                if dia_pesq and mes_pesq:
                    if mes_pesq > 6 or (mes_pesq == 6 and dia_pesq >= 28):
                        is_mata_mata = True

                query = f"'{ID_PASTA_DRIVE}' in parents and trashed = false"
                resultados = service.files().list(q=query, fields="files(id, name)").execute()
                mapa_arquivos = {arq['name']: arq['id'] for arq in resultados.get('files', [])}
                
                id_controle = mapa_arquivos['Arquivo_de_controle.xlsx']
                req_c = service.files().get_media(fileId=id_controle)
                bytes_c = io.BytesIO()
                baix_c = MediaIoBaseDownload(bytes_c, req_c)
                while not baix_c.next_chunk()[1]: pass
                bytes_c.seek(0)
                wb_leitura = openpyxl.load_workbook(bytes_c, data_only=True)
                
                jogos_reais_do_dia = {}
                
                if not is_mata_mata:
                    ws_fase = wb_leitura['FASE 1']
                    for r in range(3, 75):
                        val_data = ws_fase[f'C{r}'].value
                        if val_data is None: continue
                        data_jogo = f"{val_data.day}/{val_data.month}" if isinstance(val_data, dt.datetime) else str(val_data).strip()
                        if "/" in data_jogo and not isinstance(val_data, dt.datetime):
                            try: data_jogo = f"{int(data_jogo.split('/')[0])}/{int(data_jogo.split('/')[1])}"
                            except: pass
                        
                        if data_jogo == pesq_limpa:
                            jogos_reais_do_dia[r] = {
                                'time_casa': str(ws_fase[f'F{r}'].value).strip(),
                                'g_casa': ws_fase[f'G{r}'].value,
                                'g_fora': ws_fase[f'I{r}'].value,
                                'time_fora': str(ws_fase[f'J{r}'].value).strip()
                            }
                else:
                    nome_aba_fase2 = 'FASE 2, 3, 4, SEMI & FINAL'
                    if nome_aba_fase2 in wb_leitura.sheetnames:
                        ws_fase = wb_leitura[nome_aba_fase2]
                        for r in range(3, 100):
                            val_data = ws_fase[f'A{r}'].value
                            if val_data is None: continue
                            data_jogo = f"{val_data.day}/{val_data.month}" if isinstance(val_data, dt.datetime) else str(val_data).strip()
                            if "/" in data_jogo and not isinstance(val_data, dt.datetime):
                                try: data_jogo = f"{int(data_jogo.split('/')[0])}/{int(data_jogo.split('/')[1])}"
                                except: pass
                            
                            if data_jogo == pesq_limpa:
                                nome_casa = ws_fase[f'D{r}'].value
                                if nome_casa is not None and str(nome_casa).strip() != "":
                                    jogos_reais_do_dia[r] = {
                                        'time_casa': str(nome_casa).strip(),
                                        'g_casa': ws_fase[f'E{r}'].value,
                                        'g_fora': ws_fase[f'G{r}'].value,
                                        'time_fora': str(ws_fase[f'H{r}'].value).strip()
                                    }

                ws_tabela = wb_leitura['TABELA']
                participantes = []
                for r in range(2, 101):
                    nome = ws_tabela[f'A{r}'].value
                    if nome and str(nome).strip() != "" and str(nome).strip().lower() != "participante":
                        participantes.append(str(nome).strip())
                wb_leitura.close()
                
                dict_ranking = {p['nome']: p for p in st.session_state.ranking_processado} if st.session_state.ranking_processado else {}
                dados_painel = []
                melhor_pontuacao_dia = -1
                herois_do_dia = []
                
                palpites_por_jogo = {r: [] for r in jogos_reais_do_dia.keys()}

                for nome in participantes:
                    arq_palpite = f"{nome}.xlsx"
                    if arq_palpite in mapa_arquivos:
                        id_part = mapa_arquivos[arq_palpite]
                        palpites_usuario = carregar_palpites_em_cache(id_part)
                        
                        pontos_do_dia = 0
                        cards_html = ""
                        
                        for r, jogo_real in jogos_reais_do_dia.items():
                            palpites_fase_certa = palpites_usuario['fase2'] if is_mata_mata else palpites_usuario['fase1']
                            palp_c, palp_f = palpites_fase_certa.get(r, (None, None))
                            
                            if palp_c is not None and palp_f is not None:
                                try:
                                    pc_int, pf_int = int(float(palp_c)), int(float(palp_f))
                                    palpites_por_jogo[r].append((pc_int, pf_int))
                                except: pass
                            
                            try: p_c_str = str(int(float(palp_c))) if palp_c is not None else "-"
                            except: p_c_str = "-"
                            try: p_f_str = str(int(float(palp_f))) if palp_f is not None else "-"
                            except: p_f_str = "-"
                            
                            cor_borda = "#334155" 
                            if jogo_real['g_casa'] is not None and jogo_real['g_fora'] is not None:
                                pts = calcular_pontos(jogo_real['g_casa'], jogo_real['g_fora'], palp_c, palp_f)
                                pontos_do_dia += pts
                                if pts == 7: cor_borda = "#166534" 
                                elif pts in [2, 3, 4]: cor_borda = "#2563EB" 
                                else: cor_borda = "#991B1B" 
                                
                            cards_html += f"""
                            <div style="background:#0F172A; border:1px solid {cor_borda}; border-radius:6px; padding:8px 12px; margin-bottom:8px; display:flex; justify-content:space-between; align-items:center;">
                               <span style="color:#CBD5E1; width:30%; text-align:right; font-size:14px;">{jogo_real['time_casa']}</span>
                               <span style="color:#F8FAFC; font-weight:bold; width:40%; text-align:center; font-size:15px;">{p_c_str} x {p_f_str}</span>
                               <span style="color:#CBD5E1; width:30%; text-align:left; font-size:14px;">{jogo_real['time_fora']}</span>
                            </div>
                            """
                        
                        info = dict_ranking.get(nome, {'pontos': '?', 'posicao': '-'})
                        pos_str = str(info['posicao']).replace("º Lugar", "").strip()
                        
                        emoji = "👤"
                        if pos_str == '1': emoji = "🥇"
                        elif pos_str == '2': emoji = "🥈"
                        elif pos_str == '3': emoji = "🥉"
                        
                        cabecalho = f"{emoji} {nome} | {info['pontos']} pts | {info['posicao']}"
                        dados_painel.append({'cabecalho': cabecalho, 'cards_html': cards_html, 'pontos_dia': pontos_do_dia})
                        
                        if pontos_do_dia > melhor_pontuacao_dia:
                            melhor_pontuacao_dia = pontos_do_dia
                            herois_do_dia = [nome]
                        elif pontos_do_dia == melhor_pontuacao_dia and pontos_do_dia > 0:
                            herois_do_dia.append(nome)

            except Exception as e:
                st.error(f"Ocorreu um erro na busca: {e}")

        if st.session_state.busca_ativa:
            if modo_grafico:
                st.divider()
                st.markdown("## 📊 Relatório Oficial da Comunidade")
                
                if not jogos_reais_do_dia:
                    st.warning("Nenhum jogo encontrado para esta data.")
                
                for r, jogo in jogos_reais_do_dia.items():
                    palpites_jogo = palpites_por_jogo[r]
                    if not palpites_jogo: continue
                    
                    N = len(palpites_jogo)
                    t_c = jogo['time_casa'].upper()
                    t_f = jogo['time_fora'].upper()
                    
                    v_casa = sum(1 for c, f in palpites_jogo if c > f)
                    v_emp  = sum(1 for c, f in palpites_jogo if c == f)
                    v_fora = sum(1 for c, f in palpites_jogo if c < f)
                    
                    pct_c = int(round((v_casa / N) * 100))
                    pct_e = int(round((v_emp / N) * 100))
                    pct_f = int(round((v_fora / N) * 100))
                    
                    bar_c = "█" * int((pct_c / 100) * 10)
                    bar_e = "█" * int((pct_e / 100) * 10)
                    bar_f = "█" * int((pct_f / 100) * 10)
                    
                    placares_str = [f"{c}x{f}" for c, f in palpites_jogo]
                    top_placares = Counter(placares_str).most_common(5)
                    max_placar_count = top_placares[0][1] if top_placares else 1
                    placar_oficial = top_placares[0][0].replace("x", " x ") if top_placares else "0 x 0"
                    
                    med_c = sum(c for c, f in palpites_jogo) / N
                    med_f = sum(f for c, f in palpites_jogo) / N
                    max_med = max(med_c, med_f, 1)
                    bar_mc = "█" * int((med_c / max_med) * 8)
                    bar_mf = "█" * int((med_f / max_med) * 8)
                    
                    html_card = f"<div style='background:#0F172A; border:1px solid #1E293B; border-radius:12px; padding:20px; margin-bottom:25px; color:#F8FAFC; font-family: sans-serif; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.5);'>"
                    
                    html_card += f"<h3 style='text-align:center; color:#E2E8F0; letter-spacing: 2px; margin-top:0; margin-bottom:10px; font-size:18px;'>{t_c} x {t_f}</h3>"
                    html_card += "<p style='color:#94A3B8; font-size:12px; text-transform:uppercase; text-align:center; margin-bottom:5px;'>🏆 Palpite oficial do bolão</p>"
                    html_card += f"<p style='font-size:22px; font-weight:bold; color:#F8FAFC; text-align:center; margin:0 0 15px 0;'>{jogo['time_casa']} <span style='color:#F59E0B;'>{placar_oficial}</span> {jogo['time_fora']}</p>"
                    html_card += "<hr style='border: 0; border-top: 1px solid #1E293B; margin: 15px 0 20px 0;'>"
                    
                    html_card += "<div style='display: flex; flex-wrap: wrap; justify-content: space-between; gap: 20px;'>"
                    
                    html_card += "<div style='flex: 1; min-width: 200px;'>"
                    html_card += "<p style='color:#94A3B8; font-size:12px; text-transform:uppercase; margin-bottom:15px; border-bottom: 1px solid #1E293B; padding-bottom:5px;'>Quem vence?</p>"
                    html_card += f"<div style='display: flex; margin-bottom: 8px; font-size:14px;'><div style='width: 50px; text-align: right; margin-right: 12px;'>{jogo['time_casa']}</div><div style='flex-grow: 1; color:#22C55E;'>{bar_c}</div><div style='width: 45px; text-align: right; margin-left: 8px;'>{pct_c}%</div></div>"
                    html_card += f"<div style='display: flex; margin-bottom: 8px; font-size:14px;'><div style='width: 50px; text-align: right; margin-right: 12px;'>Empate</div><div style='flex-grow: 1; color:#64748B;'>{bar_e}</div><div style='width: 45px; text-align: right; margin-left: 8px;'>{pct_e}%</div></div>"
                    html_card += f"<div style='display: flex; margin-bottom: 8px; font-size:14px;'><div style='width: 50px; text-align: right; margin-right: 12px;'>{jogo['time_fora']}</div><div style='flex-grow: 1; color:#3B82F6;'>{bar_f}</div><div style='width: 45px; text-align: right; margin-left: 8px;'>{pct_f}%</div></div>"
                    html_card += "</div>"
                    
                    html_card += "<div style='flex: 1; min-width: 200px;'>"
                    html_card += "<p style='color:#94A3B8; font-size:12px; text-transform:uppercase; margin-bottom:15px; border-bottom: 1px solid #1E293B; padding-bottom:5px;'>Placares mais apostados</p>"
                    for placar, count in top_placares:
                        bar_len = int((count / max_placar_count) * 8)
                        bar_p = "█" * max(1, bar_len)
                        html_card += f"<div style='display: flex; margin-bottom: 8px;'><div style='width: 40px; text-align: right; margin-right: 12px; color:#F8FAFC;'>{placar}</div><div style='flex-grow: 1; color:#F59E0B;'>{bar_p}</div><div style='width: 30px; text-align: left; color:#94A3B8; margin-left: 8px;'>{count}</div></div>"
                    html_card += "</div>"
                    
                    html_card += "<div style='flex: 1; min-width: 200px;'>"
                    html_card += "<p style='color:#94A3B8; font-size:12px; text-transform:uppercase; margin-bottom:15px; border-bottom: 1px solid #1E293B; padding-bottom:5px;'>Média de gols</p>"
                    html_card += f"<div style='display: flex; margin-bottom: 8px; font-size:14px;'><div style='width: 50px; text-align: right; margin-right: 12px;'>{jogo['time_casa']}</div><div style='flex-grow: 1; color:#94A3B8;'>{bar_mc}</div><div style='width: 30px; text-align: left; font-weight:bold; margin-left: 8px;'>{med_c:.1f}</div></div>"
                    html_card += f"<div style='display: flex; margin-bottom: 8px; font-size:14px;'><div style='width: 50px; text-align: right; margin-right: 12px;'>{jogo['time_fora']}</div><div style='flex-grow: 1; color:#94A3B8;'>{bar_mf}</div><div style='width: 30px; text-align: left; font-weight:bold; margin-left: 8px;'>{med_f:.1f}</div></div>"
                    html_card += "</div>"

                    html_card += "</div></div>"
                    
                    st.markdown(html_card, unsafe_allow_html=True)
            
            else:
                if jogos_reais_do_dia:
                    st.markdown("### 🏟️ Resultados Oficiais")
                    cols_reais = st.columns(len(jogos_reais_do_dia))
                    for idx, (linha, jogo) in enumerate(jogos_reais_do_dia.items()):
                        gc = int(float(jogo['g_casa'])) if jogo['g_casa'] is not None else "-"
                        gf = int(float(jogo['g_fora'])) if jogo['g_fora'] is not None else "-"
                        with cols_reais[idx % len(cols_reais)]:
                            st.markdown(f"""
                            <div style="background:#1E293B; border:1px solid #334155; border-radius:8px; padding:12px; text-align:center; margin-bottom:15px;">
                                <div style="font-size: 16px; font-weight: bold; color: #F8FAFC;">
                                    {jogo['time_casa']} &nbsp;<span style="color:#60A5FA;">{gc} x {gf}</span>&nbsp; {jogo['time_fora']}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                if jogos_reais_do_dia and melhor_pontuacao_dia > 0:
                    st.markdown("---")
                    st.markdown("### 🔥 Destaque da Rodada")
                    st.markdown(f"""
                    <div style="background: linear-gradient(90deg, #1E293B 0%, #0F172A 100%); border-left: 4px solid #F59E0B; padding: 15px; border-radius: 6px; margin-bottom: 25px;">
                        <h4 style="color: #FBBF24; margin: 0 0 5px 0;">{", ".join(herois_do_dia)}</h4>
                        <p style="color: #F8FAFC; margin: 0; font-size: 14px;">Mito da rodada somando impressionantes <strong>{melhor_pontuacao_dia} pontos</strong> só nos jogos de hoje!</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("### 📋 Palpites da Galera")
                cols_grid = st.columns(3)
                for idx, painel in enumerate(dados_painel):
                    with cols_grid[idx % 3]:
                        with st.expander(painel['cabecalho']): st.markdown(painel['cards_html'], unsafe_allow_html=True)
